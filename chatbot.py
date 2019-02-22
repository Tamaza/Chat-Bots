# necessary imports
from vk_api import VkApi
from vk_api.upload import VkUpload
from vk_api.keyboard import VkKeyboard
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
from sql import User
from time import sleep
from openpyxl import Workbook
from datetime import datetime
from traceback import format_exc
import os
import sys
import settings

# longpoll issue fix
def try_repeat(func):
    def wrapper(*args, **kwargs):
        while True:
            try:
                return func(*args, **kwargs)
            except:
                format_exc()
                sleep(15)
    return wrapper

def save_excel():
    wb = Workbook()
    ws = wb.active

    users = User.select().where(User.request_contact != '')

    ws['A1'] = "ID"
    ws['B1'] = "Name"
    ws['C1'] = "Phone Number"
    ws['D1'] = "Gender"
    ws['E1'] = "Contact"
    ws['F1'] = "Did some magic"
    ws['G1'] = "Set an appointment"

    for i, user in enumerate(users):
        ws['A' + str(2 + i)] = user.user_id
        ws['B' + str(2 + i)] = user.request_name
        ws['C' + str(2 + i)] = user.request_phone_number
        ws['D' + str(2 + i)] = user.request_gender
        ws['E' + str(2 + i)] = user.request_contact
        ws['F' + str(2 + i)] = user.request_meaning
        ws['G' + str(2 + i)] = user.request_record

    wb.save("upload.xlsx")



# getting user info
def check_user(user_id, api):
    user = User.select().where(User.user_id == user_id).first()   
    if not user:
        user = User(user_id=user_id)
        user.save()
    user_info = api.users.get(user_ids=user.user_id, fields='city, bdate')[0]

    return user, user_info

# getting words from keyboard
def to_kb(kb):
    keyboard = VkKeyboard(one_time=True)
    kb = kb.split('\n')
    for i, k in enumerate(kb):
        kb[i] = k.split(';')
    if len(kb) == 1:
        for k in kb[0]:
            keyboard.add_button(k)
        return keyboard.get_keyboard()
    if len(kb) > 1:
        for k in kb:
            for p in k:
                keyboard.add_button(p)
            if k != kb[-1]:
                keyboard.add_line()
        return keyboard.get_keyboard()

# function for message checking
def check_message(event, api, vk_session):
    # getting the text
    text = event.obj.text
    # getting user's info
    user, user_info = check_user(event.obj.from_id, api)
    if user.user_id in settings.ADMINS.split(',') and text.lower() == "!questionnaire":
        save_excel()
        upload = VkUpload(vk_session)
        doc = upload.document_message(doc="upload.xlsx", title="upload.xlsx", peer_id=user.user_id)
        time = datetime.strftime(datetime.now(), '[%d.%m.%Y]')
        attachment = "doc" + str(doc[0]['owner_id']) + "_" + str(doc[0]['id'])
        api.messages.send(user_id=user.user_id, message=f"questionnaire{time}", attachment=attachment)
        return True
    elif user.user_id in settings.ADMINS.split(',') and text.lower() == "!restart":
        message = f"restarting the bot..."
        api.messages.send(user_id=user.user_id, message=message)
        os.execl(sys.executable,*([sys.executable]+sys.argv))
    # checking which stage are we at 
    elif user.level == "0,0,0":
        # checking the start option
        if text.lower() == "start":
            # the message that will be shown after start
            message = f"🤖Hello {user_info['first_name']}, if you remember me, I'm bot Robert."
            
            api.messages.send(user_id=user.user_id, message=message)
           
            message = "I've been waiting for this for a long time, I'm going to show you our quick demo.😊"
            api.messages.send(user_id=user.user_id, message=message)
          
            message = "write "continue" and let's dive in!"
            api.messages.send(user_id=user.user_id, message=message)
            # moving to next stage of our demo
            user.level = "1,0,0"
           
            user.save()
           
            return True
        else:
            message = "🤖 write "start" and let's go!"
            api.messages.send(user_id=user.user_id, message=message)
            return True
    elif user.level == "1,0,0":
        if text.lower() == "continue":
            message = "🤖Begin. 💌You will be able to create mailings with questionnaires and to send them to all users of the same group to remind of the services and actions and also to collect data of clients."
            api.messages.send(user_id=user.user_id, message=message)
            
            message = "👉Press the button "Example!" or send any message. in the Demo, buttons of VKontakte are used. You can deploy them, having pressed the special badge located in the field for input of the message. 😉 you should update the mobile application if you do not see this badge."
            
            keyboard = to_kb("Example!")
           #sending the message form keyboard
            api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)

            user.level = "1,1,0"
            user.save()
            return True
    elif user.level == "1,1,0":
        message = "👆Hey … Did not you forget about us yet? The Dumbldor company gives you a free set of the wizard ✨. Only we need to know you better,before you take it away"
        api.messages.send(user_id=user.user_id, message=message)
        message = "Gender\n1.🙇Male\n2.🙋Female"
        api.messages.send(user_id=user.user_id, message=message)

        user.level = "1,1,1"
        user.save()
        return True
    elif user.level == "1,1,1":
        if text.lower() in ["Male", "Female"]:
            user.request_gender = text.lower()
            message = "🙌what would you to if you had a magic wand?"
            api.messages.send(user_id=user.user_id, message=message)

            user.level = "1,1,2"
            user.save()
        else:
            message = "Hm... I'can't take that answer, maybe choose from these two."
            api.messages.send(user_id=user.user_id, message=message)
        return True
    elif user.level == "1,1,2":
        user.request_meaning = text

        message = f"Your answer “{user.request_gender}”, “{user.request_meaning}” was recorder, thank you! 😊"
        
        api.messages.send(user_id=user.user_id, message=message)

        message = "🤖Press the button or write any text to continue the demo"
        keyboard = to_kb("Continue")
        api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)

        user.level = "2,0,0"
        user.save()
        return True
    elif user.level == "2,0,0":
        message = "Main Menu, here you can group the info about your company and different departments"
        keyboard = to_kb("About us;FAQ\nWhy us?;Send Gif")
        api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

        user.level = "2,1,0"
        user.save()
        return True
    # Menu
    elif user.level == "2,1,0":
        if text == "About us":
            message = "😜We are young group of developers ❤we love our job!💥 we believe that bots can make your life easier, and take many routine tasks on their shoulders to make your life less compliacted😊"
            keyboard = to_kb("Menu;Continue demo")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
        elif text == "FAQ":
            message = "🤖I combined the questions that you might be interested in!"
            keyboard = to_kb("Robert, do you have any feelings at all?\nRobert,where are you at?\nMenu\nContinue demo")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,2"
            user.save()
            return True
        elif text == "Why us?":
            message = "💪Why not?... Ok, frankly, we are young group of developers, we treat every new project as a new adventure, we will put maximum effort to make you feel satisfied with your product."
            keyboard = to_kb("Menu;Continue demo")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
        elif text == "Send Gif":
            attachment = "doc454025337_486021163"
            keyboard = to_kb("Menu;Continue demo")
            api.messages.send(user_id=user.user_id, attachment=attachment, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
        else:
            message = "Sorry! there no such option in the Menu"
            keyboard = to_kb("About us;FAQ\nWhy us?;Send Gif")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,0"
            user.save()
            return True
    #Why us and about us
    elif user.level == "2,1,1":
        if text == "Menu":
            message = "Main menu, here you can group info about the company and its departments"
            keyboard = to_kb("About us;FAQ\nWhy us?;Send Gif")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,0"
            user.save()
            return True
        elif text == "Continue Menu":
            message = "🤖You can even set up a meeting with clients!"
            keyboard = to_kb("Example")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "3,0,0"
            user.save()
            return True
        else:
            message = "Hm... it seems you made a mistake"
            keyboard = to_kb("Menu;Continue demo")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,1"
            user.save()
            return True
    # FAQ
    elif user.level == "2,1,2":
        if text == "Robert, do you have any feeling at all?":
            message = "🤖ehh... my creaters have never thought of doing it, but I obviously like you :)"
            keyboard = to_kb("Robert,do you have any feelings at all?\nRobert whre are you at?\nMenu\nContinue demo")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)
            return True
        elif text == "Robert, what's your life is like?":
            message = "🤖Talking to clients,you know,living on server,vacations, good salary, living the dream"
            keyboard = to_kb("Do you have any feelings at all?\nRobert where are you at?\nMenu\nContinue")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)
            return True
        elif text == "Menu":
            message = "Main menu, here you can group info about the company and its departments"
            keyboard = to_kb("About us;FAQ\nWhy us?;Send Gif")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,0"
            user.save()
            return True
        elif text == "Continue demo":
            message = "🤖you can even set up an appointment with clients!"
            keyboard = to_kb("Example")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "3,0,0"
            user.save()
            return True
        else: 
            message = "Hmm... it seems like you made a mistake"
            keyboard = to_kb("Robert, do you have any feelings?\nRobert,where are you at?\nMenu\nContinue demo")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "2,1,2"
            user.save()
            return True
    # Запись на приём
    elif user.level == "3,0,0":
        message = f"👋Hello friend, {user_info['first_name']}! come visit us"
        keyboard = to_kb("Set an appointment")
        api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

        user.level = "3,1,0"
        user.save()
        return True

    elif user.level == "3,1,0":
        message = "📞give us your phone number in fortmat of  +79181001010"
        api.messages.send(user_id=user.user_id, message=message)

        user.level = "3,1,1"
        user.save()
        return True
    elif user.level == "3,1,1":
        if text[0] == "+":
            text = text[1:]
        try:
            number = int(text)
            if len(text) == 11 and number > 0:
                message = "🌈day of the week"
                keyboard = to_kb("Mon;Tue;Wen\nThs;Fri")
                api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

                user.request_phone_number = text
                user.level = "3,1,2"
                user.save()
                return True
            else:
                message = "phone number was intered in a wrong format"
                print(text)
                api.messages.send(user_id=user.user_id, message=message)
                return True
        except ValueError:
            message = "phone number was intered in a wrong format"
            print(text)
            api.messages.send(user_id=user.user_id, message=message)
            return True

    elif user.level == "3,1,2":
        if text in ["Mon", "Tue", "Wen", "Th", "Fri"]:
            message = "last step, choose the time of the appointment(write the time in the format of 00:00)"
            api.messages.send(user_id=user.user_id, message=message)

            user.request_record = text + ","
            user.level = "3,1,3"
            user.save()
            return True
        else:
            message = "There is no day like that, choose from the options."
            keyboard = to_kb("Mon;Tue;Wen\nTh;Fri")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)
            return True
    elif user.level == "3,1,3":
        if text[0] in ["0","1","2"]:
            try:
                hours = int(text[0:2])
                minutes = int(text[3:5])
                if text[2] == ":" and len(text) == 5 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                    user.request_record += str(hours) + ":" + str(minutes)
                    day = user.request_record.split(",")[0]
                    time = text
                    message = f"Great ! We are waiting for you on {day}, {time}. We will remind you about the appointment in our VK messages😊"
                    api.messages.send(user_id=user.user_id, message=message)

                    message = "🤖Make sure to know that thi is just a small demo, bot's functionality is limited only by your imagination and resources of the messenger🌈"
                    keyboard = keyboard = to_kb("See the force!")
                    api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)
                    user.level = "3,1,4"
                    user.save()
                    return True
                else:
                    message = "Wrong day format."
                    api.messages.send(user_id=user.user_id, message=message)
                    return True
            except ValueError:
                message = "Wrong day format."
                api.messages.send(user_id=user.user_id, message=message)
                return True
            except IndexError:
                message = "Wrong day format."
                api.messages.send(user_id=user.user_id, message=message)
                return True
        else:
            try:
                hours = int("0" + text[0])
                minutes = int(text[2:4])
                if text[1] == ":" and len(text) == 4 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                    user.request_record += str(hours) + ":" + str(minutes)
                    day = user.request_record.split(",")[0]
                    time = text
                    message = f"Great ! We are waiting for you on {day}, {time}. we will remind you about the appointmend in our VK messages😊"
                    api.messages.send(user_id=user.user_id, message=message)

                    message = "🤖Make sure to know that thi is just a small demo, bot's functionality is limited only by your imagination and resources of the messenger🌈"
                    keyboard = keyboard = to_kb("See the force")
                    api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)
                    user.level = "3,1,4"
                    user.save()
                    return True
                else:
                    message = "Time is given in a wrong format."
                    api.messages.send(user_id=user.user_id, message=message)
                    return True
            except ValueError:
                message = "Time is given in a wrong format"
                api.messages.send(user_id=user.user_id, message=message)
                return True
            except IndexError:
                message = "Time is given in a wrong format."
                api.messages.send(user_id=user.user_id, message=message)
                return True
            
    elif user.level == "3,1,4":
            message = "😇Possible functionality:"
            api.messages.send(user_id=user.user_id, message=message)

            message = "🤖Chat-bot can implement the CRM system"
            api.messages.send(user_id=user.user_id, message=message)

            message = "✅Can make the payment"
            api.messages.send(user_id=user.user_id, message=message)

            message = "💨Can gather the clients info in Excel sheet"
            api.messages.send(user_id=user.user_id, message=message)

            message = "👀Search the info through your resources"
            api.messages.send(user_id=user.user_id, message=message)

            message = "💬 Communicate with third-party servers by means of API, provide huge amount of various functionality. For example, the bot will translate the message into other language or will show balance on personal account of the client"
            api.messages.send(user_id=user.user_id, message=message)

            message = "🔎 Will manage to find and compare keywords in the long message of the client and to provide the prepared answer according to these keywords"
            keyboard = to_kb("Learn more!")
            api.messages.send(user_id=user.user_id, message=message, keyboard = keyboard)

            user.level = "3,1,5"
            user.save()
            return True

    elif user.level == "3,1,5":
            message = "🌈You will be able to update scripts through a chat bot🔥"
            api.messages.send(user_id=user.user_id, message=message)

            message = "🚀You will be able to collect statistics of interaction of clients with yours a chat bot - into what sections clients come most often, dialogue duration and so on"
            api.messages.send(user_id=user.user_id, message=message)

            message = "and more interesting stuff;)"
            api.messages.send(user_id=user.user_id, message=message)

            message = "💪And of course we write a chat bots for telegram, facebook messenger, viber, and for websites."
            keyboard = to_kb("Получить подарок!")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "4,0,0"
            user.save()
            return True
            

    elif user.level == "4,0,0":
            message = f"😅Listen, {user_info['first_name']} You did a great job! We will provide the first month of technical support for your  chat-bot free of charge, in case if it is necessary to correct something or change! 🌝 Specify the phone number and the best way to contact you"
            api.messages.send(user_id=user.user_id, message=message)

            message = f"Is thir your phone number? - +{user.request_phone_number}"
            keyboard = to_kb("Yes;No")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.request_name = user_info['first_name']
            user.level = "4,1,1"
            user.save()
            return True
    elif user.level == "4,1,1":
        if text == "Yes":
            message = "😌What is the best way ro reach you? (Viber, WhatsApp, Telegram)"
            keyboard = to_kb("Viber;WhatsApp\nTelegram")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "4,2,0"
            user.save()
            return True
        elif text == "No":
            message = "Your phone number?"
            api.messages.send(user_id=user.user_id, message=message)

            user.level = "4,1,2"
            user.save()
            return True
    elif user.level == "4,1,2":
        if text[0] == "+":
            text = text[1:]
        try:
            number = int(text)
            if len(text) == 11:
                message = "😌What is the best way ro reach you? (Viber, WhatsApp, Telegram)"
                keyboard = to_kb("Viber;WhatsApp\nTelegram")
                api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

                user.level = "4,2,0"
                user.save()
                return True
            else:
                message = "Wrong number format"
                print(text)
                api.messages.send(user_id=user.user_id, message=message)
                return True
        except ValueError:
            message = "Wrong number format"
            print(text)
            api.messages.send(user_id=user.user_id, message=message)
            return True

    elif user.level == "4,2,0":
        message = '🔥Great! We will get in touch soon. Write “Restart", if you want to start over;)'
        api.messages.send(user_id=user.user_id, message=message)

        user.request_contact = text
        user.level = "4,3,0"
        user.save()
        return True
    elif user.level == "4,3,0":
        if text.lower() == "Restart":
            message = "Restarting....."
            keyboard = to_kb("Start")
            api.messages.send(user_id=user.user_id, message=message, keyboard=keyboard)

            user.level = "0,0,0"
            user.save()

            return True
        
    return False

@try_repeat
def start():
    vk_session = VkApi(token=settings.BOT_TOKEN)
    api = vk_session.get_api()

    longpoll = VkBotLongPoll(vk_session, settings.GROUP_ID)

    for event in longpoll.listen():
        if event.type == VkBotEventType.MESSAGE_NEW:
            if check_message(event, api, vk_session) == False:
                message = "Hm...it seems like you made a mistake."
                api.messages.send(user_id=event.obj.from_id, message=message)




if __name__ == "__main__":
    start()
