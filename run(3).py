# -*- coding: utf-8 -*-

# Делаем необходимые ипорты
import json
import random
from flask import Flask, request
from pymessenger2.bot import Bot

import sys
from sql import User
from time import sleep
from openpyxl import Workbook
from datetime import datetime
from traceback import format_exc
import settings

# фикс бага с longpoll
def try_repeat(func):
    def wrapper(*args, **kwargs):
        while True:
            try:
                return func(*args, **kwargs)
            except:
                format_exc()
                sleep(5)
    return wrapper

def save_excel():
    wb = Workbook()
    ws = wb.active

    users = User.select().where(User.request_contact != '')

    ws['A1'] = "ID"
    ws['B1'] = "Имя"
    ws['C1'] = "Номер телефона"
    ws['D1'] = "Пол"
    ws['E1'] = "Контакты"
    ws['F1'] = "Смысл жизни"
    ws['G1'] = "Запись на приём"

    for i, user in enumerate(users):
        ws['A' + str(2 + i)] = user.user_id
        ws['B' + str(2 + i)] = user.request_name
        ws['C' + str(2 + i)] = user.request_phone_number
        ws['D' + str(2 + i)] = user.request_gender
        ws['E' + str(2 + i)] = user.request_contact
        ws['F' + str(2 + i)] = user.request_meaning
        ws['G' + str(2 + i)] = user.request_record

    wb.save("upload.xlsx")



# Функция для получения информации о пользователе
def check_user(user_id, api):
    user = User.select().where(User.user_id == user_id).first()
    if not user:
        user = User(user_id=user_id)
        user.save()
    user_info = api.get_user_info(user_id)

    return user, user_info

def start():
    app = Flask(__name__)
    app.config['JSON_AS_ASCII'] = False

    VERIFY_TOKEN = settings.VERIFY_TOKEN

    bot = Bot(settings.BOT_TOKEN)


    @app.route('/', methods=['GET', 'POST'])
    def receive_message():
        if request.method == 'GET':
            token_sent = request.args['hub.verify_token']
            return verify_fb_token(token_sent)
        else:
            output = json.loads(request.get_data().decode('utf-8'))
            for event in output['entry']:
                messaging = event['messaging']
                for message in messaging:
                    if message.get('message') or message.get('postback'):
                        user_id = message['sender']['id']
                        if message.get('postback'):
                            if check_message(message['postback']['title'], user_id) == None:
                                send_message(user_id, "Hmm...It seems like you made a mistake.")
                        elif message['message'].get('text'):
                            if check_message(message['message']['text'], user_id) == None:
                                send_message(user_id, "Hmm...It seems like you made a mistake.")
            return "Message Processed"

    def check_message(text, user_id):
        user, user_info = check_user(user_id, bot)
        print(user_info['id'])

        if user.user_id in settings.ADMINS.split(',') and text.lower() == "!анкета":
            save_excel()
            time = datetime.strftime(datetime.now(), '[%d.%m.%Y]')
            send_message(user_id, f"Анкеты{time}")
            send_file(user_id, "upload.xlsx")
            return True
        elif user.user_id in settings.ADMINS.split(',') and text.lower() == "!рестарт":
            message = f"Перезапускаем бота..."
            send_message(user_id, message)
            os.execl(sys.executable,*([sys.executable]+sys.argv))
        # Проверяем в каком отрезке сценария он находится
        elif user.language == "":
            country = ["English", "Russian", "Georgian"]
            if text in country:
                if text == "English":
                    user.language = "en"
                    message = "🤖Write “Start” to start my old engine!"
                elif text == "Russian":
                    user.language = "ru"
                    message = "🤖Напишите “Начать” чтобы завести мой заржавелый механизм!"
                elif text == "Georgian":
                    user.language = "ge"
                    message = "🤖დაწერეთ “დაწყება” რათა ჩართოთ ჩემი დაძველებული სისტემა!"
                user.save()
                send_message(user_id, message)
            else:
                message = "Hi, there! Please, select the language of the chatbot"
                keyboard = to_kb("English;Russian;Georgian")
                send_keyboard(user_id, message, keyboard)
            return True

        elif user.language == "ge":
            if user.level == "0,0,0":
                # Проверяем соответствует ли текст строке "начать"
                if text.lower() == "დაწყება":
                    # Указываем сообщение, которое будет отправлено
                    message =f"🤖გამარჯობა{user_info['first_name']}, თუ გახსოვარ მე ბოტი ამირანი ვარ GE"
                    # Отправляем сообщение, user_id берем из базы данных, сообщение берем, которое указали ранее
                    send_message(user_id, message)
                    # Тоже самое
                    message = "მე ამ მომენტს დიდხანს ველოდი.აუცილებლად გაგატარებთ ამ პატარა დემო ვერსიას და გაჩვენებთ რა შემიძლია 😊"
                    send_message(user_id, message)
                    # Тоже самое
                    message = "დაწერეთ “დაწყება” და წავედით!"
                    send_message(user_id, message)
                    # Переход к следующему участку сценария, меняем переменную level в базе данных, чтобы понимать, где находится пользователь
                    user.level = "1,0,0"
                    # Сохраняем внесенные изменения в базу данных
                    user.save()
                    # Возвращаем True, так как сообщение подошло
                    return True
                else:
                    message = "🤖 დაწერეთ “დაწყება” რათა ჩართოთ ჩემი დაძველებული სისტემა"
                    send_message(user_id, message)
                    return True
            elif user.level == "1,0,0":
                if text.lower() == "გაგრძელება":
                    message = "🤖დავიწყოთ მარტივით 💌თქვენ შეგეძლებათ გაუგზავნოთ შეტყობინებები ჯგუფებს ან კონკრეტულ ადამიანს, რათა შეახსენოთ მათ თქვენი სერვისების შესახაბ ან მიიღოთ სასურველი ინფორმაცია."
                    send_message(user_id, message)
                    send_message(user_id, message)
                    # Тоже самое
                    message = "👉დააჭირეთ ღილაკს “მაგალითი!” ან დაწერეთ ნებისმიერი ტექსტი გაგრძელებისთვის"
                    # Объявляем клавиатуру. указывается через функцию to_kb, в значении передается строка. Кнопки клавиатуры разделяются
                    # знаком ;, если нужно, чтобы кнопка была с новой строки, ставится ;
                    keyboard = to_kb("მაგალითი")
                    # Отправляем сообщение с клавиатурой
                    send_keyboard(user_id, message, keyboard)

                    user.level = "1,1,0"
                    user.save()
                    return True
            elif user.level == "1,1,0":
                message = "👆კიდევ აქ ხარ ხომ?ხომ არ ჩამოგეძინა, მიდი უცბად შენ შესახებ ცოტა ინფორმაცია მომაწოდე და გავაგრძელოთ"
                send_message(user_id, message)
                message = "შენი სქესი\n1.🙇მამრობითი\n2.🙋მდედრობითი"
                send_message(user_id, message)

                user.level = "1,1,1"
                user.save()
                return True
            elif user.level == "1,1,1":
                if text.lower() in ["მამრობითი", "მდედრობითი"]:
                    user.request_gender = text.lower()
                    message = "🙌რას გააკეთებდი თავზე საყრელი ფული რომ გქონდეს?"
                    send_message(user_id, message)

                    user.level = "1,1,2"
                    user.save()
                else:
                    message = "ეგეთი ვარიანტი არ გვაქვს, მაგრამ ამ ორიდან ამოარჩიე"
                    send_message(user_id, message)
                return True
            elif user.level == "1,1,2":
                user.request_meaning = text

                message = f"შენი პასუხი “{user.request_gender}”, “{user.request_meaning}” ჩავიწერე, გაიხარე😊"

                send_message(user_id, message)

                message = "🤖დააჭირე ღილაკს ან დაწერე ნებისმიერი ტექსტი რათა სხვა რაღაცეებიც ნახო, მაგალითად მენიუ"
                keyboard = to_kb("გაგრძელება")
                send_keyboard(user_id, message, keyboard)

                user.level = "2,0,0"
                user.save()
                return True
            elif user.level == "2,0,0":
                message = "მთავარი მენიუ, აქ შეგიძლია გამოქვეყნო ინფორმაცია შენ ბიზნესთან დაკავშირებით"
                keyboard = to_kb("ჩვენს შესახებ;ხშირად დასმული კითხვები;რატომ ჩვენ?;გიფის გაგზავნა?")
                send_keyboard(user_id, message, keyboard)

                user.level = "2,1,0"
                user.save()
                return True
            # Меню
            elif user.level == "2,1,0":
                if text == "ჩვენს შესახებ":
                    message = "😜ჩვენ ახალგაზრდა დეველოპერთა ჯგუფი ვართ ❤ჩვენ გვიყვარს ჩვენი საქმე!💥ჩვენ გვჯერა რომ ჩატ ბოტები შეძლებენ თქვენი ცხოვრების გამარტივებას და თავიდან მოგაცილებენ რუტინულ სამუშაოს😊"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
               elif text == "ხშირად დასმული კითხვები":
                    message = "🤖მე მგონია ეს კითხვები დაგაინტერესებთ!\n(1) ამირან, რაიმე გრძნობა თუ გაგაჩნია?\n(2) ამირან, სად იმყოფები?"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,2"
                    user.save()
                    return True
                elif text == "რატომ ჩვენ?":
                    message = "💪ვინ თუ არა ჩვენ?ჩვენ ძალიან მოტივირებულები ვართ რომ თქვენ მაქსიმალურად კმაყოფილად იგრძნოთ თავი, ყველა პროექტს ვუდგებით მთელი გულის-ყურით და ყველა პროექტს განსაკუთრებულად ვუდგებით"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
                elif text == "გიფის გაგზავნა":
                    message = "..."
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_image_url(user_id, "https://psv4.userapi.com/c848120/u3688817/docs/d7/e76172a1afdc/Terminator_2.gif?extra=7Op8UaaMwvyEHtFLXR-pp4YmdRZjVVCAofWU5pGs8bJPy4pGW4ZF15ijAI6H0CZvI9EpO9ZSHOTUtgijM-DGgadf6iMnPO7X-YPuboJx3UEvSl8zTvnn2xJDYepvxuwv1okWbba_hZvzqywqOdszYA")
                    send_keyboard(user_id, message, keyboard)
                    user.level = "2,1,1"
                    user.save()
                    return True
                else:
                    message = "უკაცრავად მაგრამ, ეგეთი პუნქტი არ არსებობს ჩვენს მენიუში"
                    keyboard = to_kb("ჩვენს შესახებ;ხშირად დასმული კითხვები;რატომ ჩვენ?;გიფის გაგზავნა")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
            # О нас и Почему мы
            elif user.level == "2,1,1":
                 if text == "მენიუ:
                    message = "მთავარი მენიუ, აქ შეგიძლია გამოაქვეყნო ინფორმაცია შენ ბიზნესთან დაკავშირებით"
                    keyboard = to_kb("ჩვენს შესახებ;ხშირად დასმული კითხვები;რატომ ჩვენ?;გიფის გაგზავნა")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
               elif text == "გავაგრძელოთ დემო":
                    message = "🤖კლიენტების მიღებაზე ჩაწერაც შეგეძლებათ!"
                    keyboard = to_kb("მაგალითი")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,0,0"
                    user.save()
                    return True
                else:
                    message = "მემგონი რაღაცა შეგეშალათ"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
            # FAQ
            elif user.level == "2,1,2":
                if text == "(1)":
                     message = "🤖ჩემს შემქმნელებს მაგაზე არ უფიქრიათ მაგრამ თქვენ კარგი ტიპი ჩანხართ:)"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)
                    return True
                elif text == "(2)":
                    message = "🤖აქაც მე ვარ,იქაც მე ვარ, იმიტომ რომ სერვერზე ვარ"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)
                    return True
                 elif text == "მენიუ":
                    message = "მთავარი მენიუ, აქ შეგიძლია გამოქვეყნო ინფორმაცია შენ ბიზნესთან დაკავშირებით"
                    keyboard = to_kb("ჩვენს შესახებ;ხშირად დასმული კითხვები;რატომ ჩვენ?;გიფის გაგზავნა")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
                 elif text == "გავაგრძელოთ დემო":
                    message = "🤖კლიენტების მიღებაზე ჩაწერაც შეგეძლებათ!"
                    keyboard = to_kb("მაგალითი")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,0,0"
                    user.save()
                    return True
                else:
                     message = "მემგონი რაღცა შეგეშალათ"
                    keyboard = to_kb("მენიუ;გავაგრძელოთ დემო")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,2"
                    user.save()
                    return True
            # Запись на приём
            elif user.level == "3,0,0":
               message = f"👋მოგესალმები, {user_info['first_name']}!მოდი ჩვენთან, გავიცნოთ ერთმანეთი"
                keyboard = to_kb("მიღებაზე ჩაწერა")
                send_keyboard(user_id, message, keyboard)

                user.level = "3,1,0"
                user.save()
                return True

            elif user.level == "3,1,0":
                message = "📞შეიყვანეთ თქვენი მობილური -> +99512345678ფორმატით"
                send_message(user_id, message)

                user.level = "3,1,1"
                user.save()
                return True
            elif user.level == "3,1,1":
                if text[0] == "+":
                    text = text[1:]
                try:
                    number = int(text)
                    if len(text) == 11 and number > 0:
                        message = "🌈რომელ დღეს ისურვებდით"
                        keyboard = to_kb("ორ;სამ;ოთხ;ხუთ;პარ")
                        send_keyboard(user_id, message, keyboard)

                        user.request_phone_number = text
                        user.level = "3,1,2"
                        user.save()
                        return True
                    else:
                        message = "ტელეფონის ნომერი არასწორ ფორმატით არის შეყვანილი"
                        send_message(user_id, message)
                        return True
                except ValueError:
                     message = "ტელეფონის ნომერი არასწორ ფორმატით არის შეყვანილი"
                    send_message(user_id, message)
                    return True

            elif user.level == "3,1,2":
                if text in ["ორ", "სამ", "ოთხ", "ხუთ", "პარ"]:
                    message = "და ბოლოს,რომელ საათზე მოხვალთ(შეიყვანეთ სასურველი დრო  00:00 ფორმატით)"
                    send_message(user_id, message)

                    user.request_record = text + ","
                    user.level = "3,1,3"
                    user.save()
                    return True
                else:
                    message = "ეგეთი დღე ვერ ჩამოიტანეს, აირჩიეთ ვარიანტებიდან."
                    keyboard = to_kb("ორ;სამ;ოთხ;ხუთ;პარ")
                    send_keyboard(user_id, message, keyboard)
                    return True
            elif user.level == "3,1,3":
                if text[0] in ["0","1","2"]:
                    try:
                        hours = int(text[0:2])
                        minutes = int(text[3:5])
                        if text[2] == ":" and len(text) == 5 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                            user.request_record += str(hours) + ":" + str(minutes)
                            day = user.request_record.split(",")[0]
                            time = text
                             message = f"მშვენიერია !გელოდებით {day}, {time}. შეგახსენეთ ვიზიტის დროს ფეისბუქის საშუალებით😊"
                            send_message(user_id, message)

                            message = "🤖გაითვალისწინეთ რომ ეს მხოლოდ მოკლე დემოა, ჩემი შესაძლებლობები მხოლოდ თქვენი ფანტაზიით არის შეზღუდული(ნუ კიდე მესენჯერის შესაძლებლობებითაც)🌈"
                            keyboard = keyboard = to_kb("ნახვა იმისა თუ რისი მაქნისი ვარ")
                            send_keyboard(user_id, message, keyboard)
                            user.level = "3,1,4"
                            user.save()
                            return True
                        else:
                            message = "დრო არასწორ ფორმატით არის მითითებული"
                            send_message(user_id, message)
                            return True
                    except ValueError:
                        message = "დრო არასწორ ფორმატით არის მითითებული"
                        send_message(user_id, message)
                        return True
                    except IndexError:
                       message = "დრო არასწორ ფორმატით არის მითითებული"
                        send_message(user_id, message)
                        return True
                else:
                    try:
                        hours = int("0" + text[0])
                        minutes = int(text[2:4])
                        if text[1] == ":" and len(text) == 4 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                            user.request_record += str(hours) + ":" + str(minutes)
                            day = user.request_record.split(",")[0]
                            time = text
                            message = f"მშვენიერია !გელოდებით{day}, {time}.შეგახსენეთ ვიზიტის დროს ფეისბუქის საშუალებით😊"
                            send_message(user_id, message)

                             message = "🤖გაითვალისწინეთ რომ ეს მხოლოდ მოკლე დემოა, ჩემი შესაძლებლობები მხოლოდ თქვენი ფანტაზიით არის შეზღუდული(ნუ კიდე მესენჯერის შესაძლებლობებითაც)🌈"
                            keyboard = keyboard = to_kb("ნახვა იმისა თუ რისი მაქნისი ვარ")
                            send_keyboard(user_id, message, keyboard)
                            user.level = "3,1,4"
                            user.save()
                            return True
                        else:
                            message = "დრო არასწორ ფორმატით არის მითითებული"
                            send_message(user_id, message)
                            return True
                    except ValueError:
                        message = "დრო არასწორ ფორმატით არის მითითებული"
                        send_message(user_id, message)
                        return True
                    except IndexError:
                        message = "დრო არასწორ ფორმატით არის მითითებული"."
                        send_message(user_id, message)
                        return True

            elif user.level == "3,1,4":
                    message = "😇შესაძლო ფუნქციები:"
                    send_message(user_id, message)

                    message = "🤖ბოტს შეუძლია CRM სისტემის ინტეგრირება""
                    send_message(user_id, message)

                   message = "✅შეუძლია თანხის მიღება(ონლაინ გადახდა)"
                    send_message(user_id, message)

                    message = "💨შეუძლია კლიენტების შესახებ შეგროვებული ინფორმაცია ცალკე ექსელის ფაილში გადაიტანოს"
                    send_message(user_id, message)

                    message = "👀შეუძლია ინფორმაციის მოძიება თქვენ მონაცემთა ბაზებში"
                    send_message(user_id, message)

                   message = "💬დაუკავშირდეს უცხო სერვერებს API-ის დახმარებით, რაც მის ბევრ ფუნქციას ხნის. მაგალითად, შეუძლია ტექსტის თარგმნა ან კლიენტის ბარათზე არსებული ანგარიშის ჩვენება"
                    send_message(user_id, message)

                   message = "🔎შეუძლია კლიენტის ტექტში მოძებნოს ძირითადი სიტყვები და ამ სიტყვების მიხედვით მიაწოდოს გამზადებული პასუხი"
                    keyboard = to_kb("უფრო მეტის ფუნქციის ნახვა!")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,1,5"
                    user.save()
                    return True

            elif user.level == "3,1,5":
                   message = "🌈შეგეძლებათ სკირპტების განახლება თვითონ ჩატ-ბოტის დახმარებით🔥"
                    send_message(user_id, message)

                   message = "🚀შეგეძლებათ შეეგროვოთ ინფორმაცია კლიენტის შესახებ,მაგალითად თუ რომელ განყოფილებასი გადადის მომხმარებელი ყველაზე ხშირად, რაზე ხარჯავს მეტს დროს და ა.შ"
                    send_message(user_id, message)

                     message = "და კიდევ ბევრი საინტერესო რამ :)"
                    send_message(user_id, message)

                    message = "💪რათქმაუნდა ჩვენ ვწერთ ჩატ-ბოტებს telegram, VKontakte, viber, და ვებ-საიტებისთვის ."
                    keyboard = to_kb("Получить подарок!")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "4,0,0"
                    user.save()
                    return True


            elif user.level == "4,0,0":
                    message = f"😅ყოჩაღ, {user_info['first_name']} კარგად გაართვით თავი! ჩვენ გთავაზოთ ერთ თვიან უფასო ტექნიკურ დახმარებას,რაიმე პრობლემის შექმნის ან უბრალოდ რაიმეს შესწორების შემთხვევაში 🌝დაგვიტოვეთ თვენი ნომერი მომავალი თამამშრომლობისთვის"
                    send_message(user_id, message)

                    message = f"ეს თქვენი ტელეფონის ნომერია? - +{user.request_phone_number}"
                    keyboard = to_kb("კი;არა")
                    send_keyboard(user_id, message, keyboard)

                    user.request_name = user_info['first_name']
                    user.level = "4,1,1"
                    user.save()
                    return True
            elif user.level == "4,1,1":
                if text == "კი":
                    message = "😌რითი დაგიკავშირდეთ? (Viber, WhatsApp, Telegram)"
                    keyboard = to_kb("Viber;WhatsApp;Telegram")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "4,2,0"
                    user.save()
                    return True
                elif text == "არა":
                    message = "თქვენი ტელეფონის ნომერი?"
                    send_message(user_id, message)

                    user.level = "4,1,2"
                    user.save()
                    return True
            elif user.level == "4,1,2":
                if text[0] == "+":
                    text = text[1:]
                try:
                    number = int(text)
                    if len(text) == 11:
                        message = "😌რითი დაკავშირებით? (Viber, WhatsApp, Telegram)"
                        keyboard = to_kb("Viber;WhatsApp;Telegram")
                        send_keyboard(user_id, message, keyboard)

                        user.level = "4,2,0"
                        user.save()
                        return True
                    else:
                        message = "ტელეფონის ნომერი არასწორი ფორმატით არის მითითებული"
                        send_message(user_id, message)
                        return True
                except ValueError:
                    message = "ტელეფონის ნომერი არასწორი ფორმატით არის მითითებული"
                    send_message(user_id, message)
                    return True
            elif user.level == "4,2,0":
                message = '🔥მშვენიერია, ჩვენ აუცილებლად დაგიკავშირდეთ რათა განვიხილოთ თქვენთვის შესაძლო ბოტის გაკეთება, დაწერეთ “თავიდან დაწყება",თუ გსურთ რომ თავიდან დაიწყოთ დემო;)'
                send_message(user_id, message)

                user.request_contact = text
                user.level = "4,3,0"
                user.save()

                return True
            elif user.level == "4,3,0":
                if text.lower() == "თავიდან დაწყება":
                    message = "ვიწყებთ თავიდან... მიმდინარეობს გადატვირთვა სისტემის."
                    keyboard = to_kb("დაწყება")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "0,0,0"
                    user.save()

                    return True
        elif user.language == "en":
            if user.level == "0,0,0":
                # Проверяем соответствует ли текст строке "начать"
                if text.lower() == "start":
                    # Указываем сообщение, которое будет отправлено
                    message = f"🤖Greetings {user_info['first_name']}, if you remember me, I'm bot Robert."
                    # Отправляем сообщение, user_id берем из базы данных, сообщение берем, которое указали ранее
                    send_message(user_id, message)
                    # Тоже самое
                    message = "I've been waiting for this for a long time, I'm going to show you our quick demo. 😊"
                    send_message(user_id, message)
                    # Тоже самое
                    message = "Write “Continue”  without “ ” signs  and let's go!"
                    send_message(user_id, message)
                    # Переход к следующему участку сценария, меняем переменную level в базе данных, чтобы понимать, где находится пользователь
                    user.level = "1,0,0"
                    # Сохраняем внесенные изменения в базу данных
                    user.save()
                    # Возвращаем True, так как сообщение подошло
                    return True
                else:
                    message = "🤖Write “Start” to start my old engine!"
                    send_message(user_id, message)
                    return True
            elif user.level == "1,0,0":
                if text.lower() == "continue":
                    message = "🤖You will be able to create mailings with questionnaires and to send them to all users of the same group to offer them your services and also to collect data of clients."
                    send_message(user_id, message)
                    # Тоже самое
                    message = "👉Press the button “Example!” or send any message."
                    # Объявляем клавиатуру. указывается через функцию to_kb, в значении передается строка. Кнопки клавиатуры разделяются
                    # знаком ;, если нужно, чтобы кнопка была с новой строки, ставится ;
                    keyboard = to_kb("Example!")
                    # Отправляем сообщение с клавиатурой
                    send_keyboard(user_id, message, keyboard)

                    user.level = "1,1,0"
                    user.save()
                    return True
            elif user.level == "1,1,0":
                message = "👆Hey … Have you forgotten us already? The Dumbldor company gives you free wizard set✨. But we need to know you better before you can take it"
                send_message(user_id, message)
                message = "Write you gender-\n1.🙇Male\n2.🙋Female"
                send_message(user_id, message)

                user.level = "1,1,1"
                user.save()
                return True
            elif user.level == "1,1,1":
                if text.lower() in ["male", "female"]:
                    user.request_gender = text.lower()
                    message = "🙌what would you do if you had a magic wand?"
                    send_message(user_id, message)

                    user.level = "1,1,2"
                    user.save()
                else:
                    message = "Hm... I'can't take that answer, maybe choose from these two."
                    send_message(user_id, message)
                return True
            elif user.level == "1,1,2":
                user.request_meaning = text

                message = f"Your answer“{user.request_gender}”, “{user.request_meaning}” was recorder, thank you😊"

                send_message(user_id, message)

                message = "🤖Press the button or write any message to continue the demo and see the menu"
                keyboard = to_kb("Continue")
                send_keyboard(user_id, message, keyboard)

                user.level = "2,0,0"
                user.save()
                return True
            elif user.level == "2,0,0":
                message = "Main Menu, here you can group the information about your company, products and services in different sections."
                keyboard = to_kb("About us;FAQ\nWhy us?;Send Gif")
                send_keyboard(user_id, message, keyboard)

                user.level = "2,1,0"
                user.save()
                return True
            # Меню
            elif user.level == "2,1,0":
                if text == "About us":
                    message = "😜We are a young group of developers with a passion for creating custom chatbots!💥 we believe that bots can make your company more productive and take care of a great number of routine tasks😊"
                    keyboard = to_kb("Мenu;Continue demo")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
                elif text == "FAQ":
                    message = "🤖I combined the questions that you might be interested in!\n(1) Robert do you have any feelings at all?\n(2) Robert, where are you at?"
                    send_message(user_id, message)
                    message = "Write the number of a question that interests you."
                    keyboard = to_kb("(Menu;Continue demo")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,2"
                    user.save()
                    return True
                elif text == "Why us?":
                    message = "💪 Who else?... Ok, frankly, we are group of young developers, we treat every new project as a new adventure, we will put maximum effort to make you feel satisfied with your product."
                     keyboard = to_kb("Menu;Continue demo")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
               elif text == "Send GIF"::
                    message = "..."
                    keyboard = to_kb("Menu;Continue demo")
                    send_image_url(user_id, "https://psv4.userapi.com/c848120/u3688817/docs/d7/e76172a1afdc/Terminator_2.gif?extra=7Op8UaaMwvyEHtFLXR-pp4YmdRZjVVCAofWU5pGs8bJPy4pGW4ZF15ijAI6H0CZvI9EpO9ZSHOTUtgijM-DGgadf6iMnPO7X-YPuboJx3UEvSl8zTvnn2xJDYepvxuwv1okWbba_hZvzqywqOdszYA")
                    send_keyboard(user_id, message, keyboard)
                    user.level = "2,1,1"
                    user.save()
                    return True
                else:
                    message = "Sorry, but such option is not our menu."
                    keyboard = to_kb("About us;FAQ;Why us?;Send GIF")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
            # О нас и Почему мы
            elif user.level == "2,1,1":
               if text == "Menu":
                    message = "Main menu, here you can group info about the company and its departments"
                    keyboard = to_kb("About us;FAQ;Why us?Send GIF")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
                elif text == "Continue demo":
                    message = "🤖You can even set up a meeting with clients!"
                    keyboard = to_kb("Example")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,0,0"
                    user.save()
                    return True
                else:
                    message = "Hmm... it seems like you made a mistake"
                    keyboard = to_kb("Menu;Continue demo")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
            # FAQ
            elif user.level == "2,1,2":
                if text == "(1)":
                   message = "🤖ehh... my creaters have never thought of doing it, but I obviously like you :)"
                    keyboard = to_kb("Menu;Continue demo")
                    send_keyboard(user_id, message, keyboard)
                    return True
                elif text == "(2)":
                     message = "🤖 I'm here and there,you know, it's exciting to be at different palces simultaneously; but I have one place that I call home - The Server"
                    keyboard = to_kb("Menu;Continue demo")
                    send_keyboard(user_id, message, keyboard)
                    return True
                elif text == "Menu":
                    message = "Main menu, here you can group info about the company and its departments"
                    keyboard = to_kb("О нас;FAQ;Why us?;Send GIF")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
                elif text == "Continue demo":
                    message = "🤖 You can even set up an appointment with clients!"
                    keyboard = to_kb("Example")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,0,0"
                    user.save()
                    return True
                else:
                    message = "Hmm... it seems like you made a mistake."
                    keyboard = to_kb("Menu;Continue demo")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,2"
                    user.save()
                    return True
            # Запись на приём
            elif user.level == "3,0,0":
                 message = f"👋Well Hello, {user_info['first_name']}! let's meet each other my dude!"
                keyboard = to_kb("Set an appointment")
                send_keyboard(user_id, message, keyboard)

                user.level = "3,1,0"
                user.save()
                return True

            elif user.level == "3,1,0":
                 message = "📞 Give us your phone number in fortmat of  +79181001010"
                send_message(user_id, message)

                user.level = "3,1,1"
                user.save()
                return True
            elif user.level == "3,1,1":
                if text[0] == "+":
                    text = text[1:]
                try:
                    number = int(text)
                    if len(text) == 11 and number > 0:
                       message = "🌈 Day of the week"
                        keyboard = to_kb("Mon;Tue;Wen\nThs;Fri")
                        send_keyboard(user_id, message, keyboard)

                        user.request_phone_number = text
                        user.level = "3,1,2"
                        user.save()
                        return True
                    else:
                        message = "phone number was entered in a wrong format"
                        send_message(user_id, message)
                        return True
                except ValueError:
                    message = "phone number was entered in a wrong format"
                    send_message(user_id, message)
                    return True

            elif user.level == "3,1,2":
                if text in ["Mon", "Tue", "Wen", "Th", "Fri"]:
                    message = "Last step, choose the time of the appointment(write the time in the format of 00:00)"
                    send_message(user_id, message)

                    user.request_record = text + ","
                    user.level = "3,1,3"
                    user.save()
                    return True
                else:
                    message = "There is no day like that, choose from the options."
                    keyboard = to_kb("Mon;Tue;Wen;Th;Fri")
                    send_keyboard(user_id, message, keyboard)
                    return True
            elif user.level == "3,1,3":
                if text[0] in ["0","1","2"]:
                    try:
                        hours = int(text[0:2])
                        minutes = int(text[3:5])
                        if text[2] == ":" and len(text) == 5 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                            user.request_record += str(hours) + ":" + str(minutes)
                            day = user.request_record.split(",")[0]
                            time = text
                          message = f"Great ! we are waiting for you on {day}, {time}. we'll remind you about it through our fb messages😊"
                            send_message(user_id, message)

                           message = "🤖Make sure to know that this is just a small demo, bot's functionality is limited only by your imagination and resources of the messenger🌈"
                          keyboard = keyboard = to_kb("Unleash the power")
                            send_keyboard(user_id, message, keyboard)
                            user.level = "3,1,4"
                            user.save()
                            return True
                        else:
                            message = "Wrong time format."
                            send_message(user_id, message)
                            return True
                    except ValueError:
                         message = "Wrong time format."
                        send_message(user_id, message)
                        return True
                    except IndexError:
                        message = "Wrong time format."
                        send_message(user_id, message)
                        return True
                else:
                    try:
                        hours = int("0" + text[0])
                        minutes = int(text[2:4])
                        if text[1] == ":" and len(text) == 4 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                            user.request_record += str(hours) + ":" + str(minutes)
                            day = user.request_record.split(",")[0]
                            time = text
                           message = f"Great ! we are waiting for you on {day}, {time}. we'll remind you about it through our fb messages😊"
                            send_message(user_id, message)

                          message = "🤖Make sure to know that this is just a small demo, bot's functionality is limited only by your imagination and resources of the messenger🌈"
                            keyboard = keyboard = to_kb("Познать мощь")
                            send_keyboard(user_id, message, keyboard)
                            user.level = "3,1,4"
                            user.save()
                            return True
                        else:
                           message = "Wrong time format."
                            send_message(user_id, message)
                            return True
                    except ValueError:
                       message = "Wrong time format."
                        send_message(user_id, message)
                        return True
                    except IndexError:
                        message = "Wrong time format."
                        send_message(user_id, message)
                        return True

            elif user.level == "3,1,4":
                    message = "😇Possible functionality":
                    send_message(user_id, message)
					message = "🤖Chat-bot can implement the CRM system"
                    send_message(user_id, message)

                    message ="✅Can make the payment"
                    send_message(user_id, message)

                    message = "💨Can gather the clients info in Excel sheet"
                    send_message(user_id, message)

                    message = "👀Search the info through your resources"
                    send_message(user_id, message)

                    message = "💬 Communicate with third-party servers by means of API, provide huge amount of various functionality. For example, the bot will translate the message into other language or will show balance on personal account of the client"
                    send_message(user_id, message)

                    message = "🔎 Will manage to find and compare keywords in the long message of the client and to provide the prepared answer according to these keywords"
                    keyboard = to_kb("Feel even more power!")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,1,5"
                    user.save()
                    return True

            elif user.level == "3,1,5":
                     message =  message = "🌈You will be able to update scripts through a chat bot🔥"
                    send_message(user_id, message)

                    message = message = "🚀You will be able to collect statistics of interaction of clients with yours a chat bot - into what sections clients come most often, dialogue duration and so on"
                    send_message(user_id, message)

                    message = "and more interesting stuff :)"
                    send_message(user_id, message)

                    message = "💪And of course we write chat bots for telegram, facebook messenger, viber, and for websites."
                    keyboard = to_kb("Receive a prize")

                    user.level = "4,0,0"
                    user.save()
                    return True


            elif user.level == "4,0,0":
                     message = f"😅Listen, {user_info['first_name']}you did a great job! We will provide the first month of technical support for your  chat-bot free of charge, in case if it is necessary to correct something or change! 🌝 Specify the phone number and the best way to contact you"
                    send_message(user_id, message)

					message = f"Is this your phone number?- +{user.request_phone_number}"
                    keyboard = to_kb("Yes;No")
                    send_keyboard(user_id, message, keyboard)

                    user.request_name = user_info['first_name']
                    user.level = "4,1,1"
                    user.save()
                    return True
            elif user.level == "4,1,1":
                if text == "Yes":
                    message = "😌What is the best way ro reach you? (Viber, WhatsApp, Telegram)"
                    keyboard = to_kb("Viber;WhatsApp;Telegram")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "4,2,0"
                    user.save()
                    return True
                elif text == "No":
                    message = "Your phone number?"
                    send_message(user_id, message)

                    user.level = "4,1,2"
                    user.save()
                    return True
            elif user.level == "4,1,2":
                if text[0] == "+":
                    text = text[1:]
                try:
                    number = int(text)
                    if len(text) == 11:
                        message = "😌What is the best way ro reach you? (Viber, WhatsApp, Telegram)"
                        keyboard = to_kb("Viber;WhatsApp;Telegram")
                        send_keyboard(user_id, message, keyboard)

                        user.level = "4,2,0"
                        user.save()
                        return True
                    else:
                       message = "wrong phone number format"
                        send_message(user_id, message)
                        return True
                except ValueError:
                    message = "wrong phone number format"
                    send_message(user_id, message)
                    return True
            elif user.level == "4,2,0":
                message = '🔥Great! we will get in touch with you soon. If you want to restart the demo, type  restart)'
                send_message(user_id, message)

                user.request_contact = text
                user.level = "4,3,0"
                user.save()

                return True
            elif user.level == "4,3,0":
                if text.lower() == "restart":
                    message = "Вжух... Возвращаемся."
                    keyboard = to_kb("Начать")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "0,0,0"
                    user.save()

                    return True
        elif user.language == "ru":
            if user.level == "0,0,0":
                # Проверяем соответствует ли текст строке "начать"
                if text.lower() == "начать":
                    # Указываем сообщение, которое будет отправлено
                    message = f"🤖Приветствую {user_info['first_name']}, если помните меня - я тот самый бот Роберт."
                    # Отправляем сообщение, user_id берем из базы данных, сообщение берем, которое указали ранее
                    send_message(user_id, message)
                    # Тоже самое
                    message = "Я ждал этого момента и обязательно проведу вас по нашему короткому демо, покажу функционал и подарю кое-что в конце 😊"
                    send_message(user_id, message)
                    # Тоже самое
                    message = "Напишите “Продолжить” и поехали! Не пишите кавычки."
                    send_message(user_id, message)
                    # Переход к следующему участку сценария, меняем переменную level в базе данных, чтобы понимать, где находится пользователь
                    user.level = "1,0,0"
                    # Сохраняем внесенные изменения в базу данных
                    user.save()
                    # Возвращаем True, так как сообщение подошло
                    return True
                else:
                    message = "🤖Напишите “Начать” чтобы завести мой заржавелый механизм!"
                    send_message(user_id, message)
                    return True
            elif user.level == "1,0,0":
                if text.lower() == "продолжить":
                    message = "🤖Начнём с простого. 💌Вы сможете создавать рассылки с опросниками и отправлять их всем пользователям своей группы, чтобы напоминать о своих услугах и акциях, а также собирать данные клиентов."
                    send_message(user_id, message)
                    # Тоже самое
                    message = "👉Нажмите на кнопку “Покажи пример!” или отправьте любое сообщение. в Демо используются кнопки ВКонтакте. Вы можете развернуть их, нажав на специальный значок расположенный в поле для ввода сообщения. 😉Вам стоит обновить мобильное приложение, если сейчас вы не видите этот значок."
                    # Объявляем клавиатуру. указывается через функцию to_kb, в значении передается строка. Кнопки клавиатуры разделяются
                    # знаком ;, если нужно, чтобы кнопка была с новой строки, ставится ;
                    keyboard = to_kb("Покажи пример!")
                    # Отправляем сообщение с клавиатурой
                    send_keyboard(user_id, message, keyboard)

                    user.level = "1,1,0"
                    user.save()
                    return True
            elif user.level == "1,1,0":
                message = "👆Псс… Ты ещё не забыл о нас? Компания “Думбльдор” дарит вам бесплатный набор волшебника✨. Только нам необходимо узнать вас получше перед тем как вы заберёте его"
                send_message(user_id, message)
                message = "Ваш пол - напишите одно из слов\n1.🙇Мужской\n2.🙋Женский"
                send_message(user_id, message)

                user.level = "1,1,1"
                user.save()
                return True
            elif user.level == "1,1,1":
                if text.lower() in ["мужской", "женский"]:
                    user.request_gender = text.lower()
                    message = "🙌Что бы вы наколдовали, если у вас была волшебная палочка?"
                    send_message(user_id, message)

                    user.level = "1,1,2"
                    user.save()
                else:
                    message = "Хм, такой ответ я не приму. Выберите один вариант из двух."
                    send_message(user_id, message)
                return True
            elif user.level == "1,1,2":
                user.request_meaning = text

                message = f"Ваш ответ “{user.request_gender}”, “{user.request_meaning}” был записан, спасибо😊"

                send_message(user_id, message)

                message = "🤖Нажмите на кнопку либо отправьте любое сообщение чтобы продолжить демо и взглянуть на пример Меню"
                keyboard = to_kb("Продолжить")
                send_keyboard(user_id, message, keyboard)

                user.level = "2,0,0"
                user.save()
                return True
            elif user.level == "2,0,0":
                message = "Главное меню - здесь вы можете группировать информацию о компании и услугах в разделы."
                keyboard = to_kb("О нас;FAQ;Почему мы?;Скинуть гифку")
                send_keyboard(user_id, message, keyboard)

                user.level = "2,1,0"
                user.save()
                return True
            # Меню
            elif user.level == "2,1,0":
                if text == "О нас":
                    message = "😜Мы молодые, энергичные ребята и обожаем работать на границе между людьми и программированием. ❤Мы любим своё дело, ведь боты - это же просто чудо!💥 Мы верим в то, что чат-боты смогут взять на себя очень многие рутинные процессы, улучшить взаимодействие с клиентами, решить огромное количество проблем и сделать вас счастливым😊"
                    keyboard = to_kb("Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
                elif text == "FAQ":
                    message = "🤖Я объединил вопросы, которые не могут не интересовать вас!\n(1) Роберт, есть ли у тебя чувства?\n(2) Роберт, где ты находишься?"
                    keyboard = to_kb("(1);(2);Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,2"
                    user.save()
                    return True
                elif text == "Почему мы?":
                    message = "💪А кто же ещё? - ха-ха, ладно… Мы совсем молодые ребята, но мы любим своё дело. Нас отличает от конкурентов то, что для нас - ваш проект будет центром вселенной. Мы будем с вами абсолютно открытыми, выйдем на связь в любое время, будем гибкими и креативными."
                    keyboard = to_kb("Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
                elif text == "Скинуть гифку":
                    message = "..."
                    keyboard = to_kb("Меню;Продолжить демо")
                    send_image_url(user_id, "https://psv4.userapi.com/c848120/u3688817/docs/d7/e76172a1afdc/Terminator_2.gif?extra=7Op8UaaMwvyEHtFLXR-pp4YmdRZjVVCAofWU5pGs8bJPy4pGW4ZF15ijAI6H0CZvI9EpO9ZSHOTUtgijM-DGgadf6iMnPO7X-YPuboJx3UEvSl8zTvnn2xJDYepvxuwv1okWbba_hZvzqywqOdszYA")
                    send_keyboard(user_id, message, keyboard)
                    user.level = "2,1,1"
                    user.save()
                    return True
                else:
                    message = "Извините, но такого пункта нет в меню."
                    keyboard = to_kb("О нас;FAQ;Почему мы?;Скинуть гифку")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
            # О нас и Почему мы
            elif user.level == "2,1,1":
                if text == "Меню":
                    message = "Главное меню - здесь вы можете группировать информацию о компании и услугах в разделы."
                    keyboard = to_kb("О нас;FAQ;Почему мы?;Скинуть гифку")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
                elif text == "Продолжить демо":
                    message = "🤖Вы даже сможете записывать клиентов на приём!"
                    keyboard = to_kb("Пример")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,0,0"
                    user.save()
                    return True
                else:
                    message = "Хм, видимо вы ошиблись."
                    keyboard = to_kb("Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,1"
                    user.save()
                    return True
            # FAQ
            elif user.level == "2,1,2":
                if text == "(1)":
                    message = "🤖Эх, мои создатели не додумались до этого, но вы мне определенно нравитесь :)"
                    keyboard = to_kb("(1);(2);Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)
                    return True
                elif text == "(2)":
                    message = "🤖Я то здесь, то там… Знаешь, занимательно быть в нескольких местах одновременно. Но вообще у меня есть дом - это мой сервер"
                    keyboard = to_kb("(1);(2);Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)
                    return True
                elif text == "Меню":
                    message = "Главное меню - здесь вы можете группировать информацию о компании и услугах в разделы."
                    keyboard = to_kb("О нас;FAQ;Почему мы?;Скинуть гифку")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,0"
                    user.save()
                    return True
                elif text == "Продолжить демо":
                    message = "🤖Вы даже сможете записывать клиентов на приём!"
                    keyboard = to_kb("Пример")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,0,0"
                    user.save()
                    return True
                else:
                    message = "Хм, видимо вы ошиблись."
                    keyboard = to_kb("(1);(2);Меню;Продолжить демо")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "2,1,2"
                    user.save()
                    return True
            # Запись на приём
            elif user.level == "3,0,0":
                message = f"👋Ну здравствуй, {user_info['first_name']}! Айда к нам на приём, родненький"
                keyboard = to_kb("Записаться на прием")
                send_keyboard(user_id, message, keyboard)

                user.level = "3,1,0"
                user.save()
                return True

            elif user.level == "3,1,0":
                message = "📞Введите ваш номер телефона в формате +79181001010"
                send_message(user_id, message)

                user.level = "3,1,1"
                user.save()
                return True
            elif user.level == "3,1,1":
                if text[0] == "+":
                    text = text[1:]
                try:
                    number = int(text)
                    if len(text) == 11 and number > 0:
                        message = "🌈Удобный день недели"
                        keyboard = to_kb("Пн;Вт;Ср;Чт;Пт")
                        send_keyboard(user_id, message, keyboard)

                        user.request_phone_number = text
                        user.level = "3,1,2"
                        user.save()
                        return True
                    else:
                        message = "Номер телефона указан в неверном формате"
                        send_message(user_id, message)
                        return True
                except ValueError:
                    message = "Номер телефона указан в неверном формате"
                    send_message(user_id, message)
                    return True

            elif user.level == "3,1,2":
                if text in ["Пн", "Вт", "Ср", "Чт", "Пт"]:
                    message = "Последний шаг. Во сколько бы вы хотели придти (введите время в формате 00:00)"
                    send_message(user_id, message)

                    user.request_record = text + ","
                    user.level = "3,1,3"
                    user.save()
                    return True
                else:
                    message = "Такого дня не было в списке, выберите один из вариантов."
                    keyboard = to_kb("Пн;Вт;Ср;Чт;Пт")
                    send_keyboard(user_id, message, keyboard)
                    return True
            elif user.level == "3,1,3":
                if text[0] in ["0","1","2"]:
                    try:
                        hours = int(text[0:2])
                        minutes = int(text[3:5])
                        if text[2] == ":" and len(text) == 5 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                            user.request_record += str(hours) + ":" + str(minutes)
                            day = user.request_record.split(",")[0]
                            time = text
                            message = f"Отлично-с ! Ждём вас у нас в {day}, {time}. Мы вам напомним о приёме в сообщениях ВК😊"
                            send_message(user_id, message)

                            message = "🤖Учтите что это всего лишь скромное Демо. Возможности чат-бота ограничены только вашей фантазией и функционалом мессенджера🌈"
                            keyboard = keyboard = to_kb("Познать мощь")
                            send_keyboard(user_id, message, keyboard)
                            user.level = "3,1,4"
                            user.save()
                            return True
                        else:
                            message = "Время указано в неверном формате."
                            send_message(user_id, message)
                            return True
                    except ValueError:
                        message = "Время указано в неверном формате."
                        send_message(user_id, message)
                        return True
                    except IndexError:
                        message = "Время указано в неверном формате."
                        send_message(user_id, message)
                        return True
                else:
                    try:
                        hours = int("0" + text[0])
                        minutes = int(text[2:4])
                        if text[1] == ":" and len(text) == 4 and hours < 24 and minutes < 60 and hours >= 0 and minutes >= 0:
                            user.request_record += str(hours) + ":" + str(minutes)
                            day = user.request_record.split(",")[0]
                            time = text
                            message = f"Отлично-с ! Ждём вас у нас в {day}, {time}. Мы вам напомним о приёме в сообщениях ВК😊"
                            send_message(user_id, message)

                            message = "🤖Учтите что это всего лишь скромное Демо. Возможности чат-бота ограничены только вашей фантазией и функционалом мессенджера🌈"
                            keyboard = keyboard = to_kb("Познать мощь")
                            send_keyboard(user_id, message, keyboard)
                            user.level = "3,1,4"
                            user.save()
                            return True
                        else:
                            message = "Время указано в неверном формате."
                            send_message(user_id, message)
                            return True
                    except ValueError:
                        message = "Время указано в неверном формате."
                        send_message(user_id, message)
                        return True
                    except IndexError:
                        message = "Время указано в неверном формате."
                        send_message(user_id, message)
                        return True

            elif user.level == "3,1,4":
                    message = "😇Возможный функционал:"
                    send_message(user_id, message)

                    message = "🤖чат-бот сможет создать сделку в вашей CRM системе"
                    send_message(user_id, message)

                    message = "✅Сможет принять оплату"
                    send_message(user_id, message)

                    message = "💨Отгрузить все накопленные данные клиентов в файл Excel"
                    send_message(user_id, message)

                    message = "👀Осуществлять поиск информации по вашему ресурсу"
                    send_message(user_id, message)

                    message = "💬Общаться со сторонними серверами с помощью API, предоставляя огромное количество всевозможного функционала. Например, бот переведёт сообщение на другой язык или покажет баланс на лицевом счёте клиента"
                    send_message(user_id, message)

                    message = "🔎Сумеет найти и сопоставить ключевые слова в длинном сообщении клиента и предоставить заготовленный ответ по этим ключевым словам"
                    keyboard = to_kb("Познать больше мощи!")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "3,1,5"
                    user.save()
                    return True

            elif user.level == "3,1,5":
                    message = "🌈Вы сможете обновлять скрипты через самого чат-бота🔥"
                    send_message(user_id, message)

                    message = "🚀Вы сможете собирать статистику взаимодействия клиентов с вашим чат-ботом - в какие разделы клиенты заходят чаще всего, длительность диалога и так далее"
                    send_message(user_id, message)

                    message = "Да и много чего интересного;)"
                    send_message(user_id, message)

                    message = "💪И конечно же мы пишем чат-ботов для telegram, facebook messenger, viber, и для веб-сайтов."
                    keyboard = to_kb("Получить подарок!")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "4,0,0"
                    user.save()
                    return True


            elif user.level == "4,0,0":
                    message = f"😅Слуушайте, {user_info['first_name']} ну вы молодчина! Мы предоставим первый месяц тех. поддержки вашего чат-бота бесплатно, на случай если нужно будет что-то подправить или изменить! 🌝Уточните свой номер телефона и лучший способ связаться с вами"
                    send_message(user_id, message)

                    message = f"Это ваш номер телефона? - +{user.request_phone_number}"
                    keyboard = to_kb("Да;Нет")
                    send_keyboard(user_id, message, keyboard)

                    user.request_name = user_info['first_name']
                    user.level = "4,1,1"
                    user.save()
                    return True
            elif user.level == "4,1,1":
                if text == "Да":
                    message = "😌Как к вам лучше будет достучаться? (Viber, WhatsApp, Telegram)"
                    keyboard = to_kb("Viber;WhatsApp;Telegram")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "4,2,0"
                    user.save()
                    return True
                elif text == "Нет":
                    message = "Ваш номер телефона?"
                    send_message(user_id, message)

                    user.level = "4,1,2"
                    user.save()
                    return True
            elif user.level == "4,1,2":
                if text[0] == "+":
                    text = text[1:]
                try:
                    number = int(text)
                    if len(text) == 11:
                        message = "😌Как к вам лучше будет достучаться? (Viber, WhatsApp, Telegram)"
                        keyboard = to_kb("Viber;WhatsApp;Telegram")
                        send_keyboard(user_id, message, keyboard)

                        user.level = "4,2,0"
                        user.save()
                        return True
                    else:
                        message = "Номер телефона указан в неверном формате"
                        send_message(user_id, message)
                        return True
                except ValueError:
                    message = "Номер телефона указан в неверном формате"
                    send_message(user_id, message)
                    return True

            elif user.level == "4,2,0":
                message = '🔥Отлично! Мы с вами свяжемся совсем скоро и вместе обсудим возможности внедрения чат-бота. Напишите “Начать сначала", если хотите повторить;)'
                send_message(user_id, message)

                user.request_contact = text
                user.level = "4,3,0"
                user.save()

                return True
            elif user.level == "4,3,0":
                if text.lower() == "начать сначала":
                    message = "Вжух... Возвращаемся."
                    keyboard = to_kb("Начать")
                    send_keyboard(user_id, message, keyboard)

                    user.level = "0,0,0"
                    user.save()

                    return True

    def to_kb(keyboard):
        keyboard = keyboard.split(";")
        buttons = []
        for button in keyboard:
            buttons.append({
            "content_type":"text",
            "title": button,
            "payload": "NOTHING"
            })
        return buttons

    def get_user_info(user_id):
        return bot.get_user_info(user_id)

    def verify_fb_token(token_sent):
        if token_sent == VERIFY_TOKEN:
            return request.args['hub.challenge']
        else:
            return 'Invalid verification token'

    def send_keyboard(recipient_id, response, buttons):
        bot.send_quick_reply(recipient_id, response, buttons)
        return 'Success'

    def send_image_url(recipient_id, image_url):
        bot.send_image_url(recipient_id, image_url)
        return 'Success'

    def send_message(recipient_id, response):
        bot.send_text_message(recipient_id, response)
        return 'Success'

    def send_file(recipient_id, attachment_path):
        bot.send_file(recipient_id, attachment_path)
        return 'Success'

    def get_message():
        '''Отправляет случайные сообщения пользователю.'''
        sample_responses = ["Потрясающе!", "Я вами горжусь!", "Продолжайте в том же духе!",
                            "Лучшее, что я когда-либо видел!"]
        return random.choice(sample_responses)

    app.run(port=5002)

if __name__ == "__main__":
    start()
